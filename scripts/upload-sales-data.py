#!/usr/bin/env python3
"""
Upload sales.xlsx data to sales_history table
"""

import os
import sys
import openpyxl
import mysql.connector
from datetime import datetime
from decimal import Decimal
import uuid

# Database connection parameters
DB_HOST = os.getenv('DB_HOST', 'localhost')
DB_USER = os.getenv('DB_USER', 'root')
DB_PASSWORD = os.getenv('DB_PASSWORD', '')
DB_NAME = os.getenv('DB_NAME', 'chainmind')
DB_PORT = int(os.getenv('DB_PORT', 3306))

def parse_database_url(url):
    """Parse DATABASE_URL environment variable"""
    if not url:
        return None
    
    # Format: mysql://user:password@host:port/database?ssl=...
    try:
        # Remove protocol
        url = url.replace('mysql://', '').replace('mysql+pymysql://', '')
        
        # Remove SSL parameters
        if '?' in url:
            url = url.split('?')[0]
        
        # Split credentials and host
        if '@' in url:
            creds, host_db = url.split('@')
            user, password = creds.split(':')
        else:
            user = 'root'
            password = ''
            host_db = url
        
        # Split host and database
        if '/' in host_db:
            host_port, database = host_db.split('/', 1)
        else:
            host_port = host_db
            database = 'chainmind'
        
        # Split host and port
        if ':' in host_port:
            host, port = host_port.rsplit(':', 1)
            port = int(port)
        else:
            host = host_port
            port = 3306
        
        return {
            'host': host,
            'user': user,
            'password': password,
            'database': database,
            'port': port
        }
    except Exception as e:
        print(f"Error parsing DATABASE_URL: {e}")
        return None

def get_db_config():
    """Get database configuration from environment"""
    # Try to parse DATABASE_URL first
    db_url = os.getenv('DATABASE_URL')
    if db_url:
        config = parse_database_url(db_url)
        if config:
            return config
    
    # Fall back to individual environment variables
    return {
        'host': os.getenv('DB_HOST', 'localhost'),
        'user': os.getenv('DB_USER', 'root'),
        'password': os.getenv('DB_PASSWORD', ''),
        'database': os.getenv('DB_NAME', 'chainmind'),
        'port': int(os.getenv('DB_PORT', 3306))
    }

def generate_id(prefix='sh'):
    """Generate unique ID"""
    return f"{prefix}-{uuid.uuid4().hex[:20].upper()}"

def format_month(date_obj):
    """Format datetime to YYYY-MM format"""
    if isinstance(date_obj, datetime):
        return date_obj.strftime('%Y-%m')
    return str(date_obj)

def upload_sales_data(file_path):
    """Upload sales data from Excel to database"""
    
    # Get database config
    config = get_db_config()
    print(f"Connecting to database: {config['host']}:{config['port']}/{config['database']}")
    
    try:
        # Connect to database
        conn = mysql.connector.connect(
            host=config['host'],
            user=config['user'],
            password=config['password'],
            database=config['database'],
            port=config['port'],
            ssl_disabled=False,
            use_pure=True
        )
        cursor = conn.cursor()
        
        # Load Excel file
        print(f"Loading Excel file: {file_path}")
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Get headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        print(f"Headers: {headers}")
        
        # Prepare insert statement
        insert_sql = """
            INSERT INTO sales_history (
                history_id, month, fg_code, fg_description, division, country, channel,
                units_sold, gross_sales_aed, promo_discount_percent, net_asp_aed,
                net_sales_aed, returns_units, fill_rate_percent, trade_spend_aed,
                sell_through_percent, createdAt
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW()
            )
        """
        
        # Read and insert data
        row_count = 0
        error_count = 0
        batch_size = 100
        batch_data = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            try:
                # Map Excel columns to database fields
                month = format_month(row[0])  # Month
                fg_code = row[1]  # FG Code
                fg_description = row[2]  # FG Description
                division = row[3]  # Division
                country = row[4]  # Country
                channel = row[5]  # Channel
                units_sold = float(row[6]) if row[6] else 0  # Units Sold
                gross_sales_aed = Decimal(str(row[7])) if row[7] else Decimal('0')  # Gross Sales AED
                promo_discount_percent = float(row[8]) if row[8] else 0  # Promo Discount %
                net_asp_aed = Decimal(str(row[9])) if row[9] else Decimal('0')  # Net ASP AED
                net_sales_aed = Decimal(str(row[10])) if row[10] else Decimal('0')  # Net Sales AED
                returns_units = float(row[11]) if row[11] else 0  # Returns Units
                fill_rate_percent = float(row[12]) if row[12] else 0  # Fill Rate %
                trade_spend_aed = Decimal(str(row[13])) if row[13] else Decimal('0')  # Trade Spend AED
                sell_through_percent = float(row[14]) if row[14] else 0  # Sell Through %
                
                # Generate unique ID
                history_id = generate_id('sh')
                
                # Add to batch
                batch_data.append((
                    history_id, month, fg_code, fg_description, division, country, channel,
                    units_sold, gross_sales_aed, promo_discount_percent, net_asp_aed,
                    net_sales_aed, returns_units, fill_rate_percent, trade_spend_aed,
                    sell_through_percent
                ))
                
                # Execute batch insert
                if len(batch_data) >= batch_size:
                    cursor.executemany(insert_sql, batch_data)
                    conn.commit()
                    row_count += len(batch_data)
                    print(f"Inserted {row_count} rows...")
                    batch_data = []
                
            except Exception as e:
                error_count += 1
                print(f"Error on row {row_idx}: {e}")
                if error_count > 10:
                    print("Too many errors, stopping...")
                    break
        
        # Insert remaining batch
        if batch_data:
            cursor.executemany(insert_sql, batch_data)
            conn.commit()
            row_count += len(batch_data)
        
        # Close connections
        cursor.close()
        conn.close()
        
        print(f"\n✅ Upload complete!")
        print(f"Total rows inserted: {row_count}")
        print(f"Total errors: {error_count}")
        
        return row_count
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 0

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python upload-sales-data.py <path-to-sales.xlsx>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        sys.exit(1)
    
    upload_sales_data(file_path)
